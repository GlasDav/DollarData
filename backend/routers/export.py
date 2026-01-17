from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
import csv
import io
import json
from datetime import datetime

from .. import models, schemas
from ..database import get_db
from ..auth import get_current_user

router = APIRouter(
    prefix="/export",
    tags=["export"]
)

@router.get("/transactions")
def export_transactions(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = Query("csv", enum=["csv", "json"]),
    db: Session = Depends(get_db)
):
    # Base query
    query = db.query(models.Transaction)
    
    # Filter by date if provided
    if start_date:
        query = query.filter(models.Transaction.date >= start_date)
    if end_date:
        query = query.filter(models.Transaction.date <= end_date)
        
    transactions = query.order_by(models.Transaction.date.desc()).all()
    
    if format == "json":
        data = [
            {
                "date": t.date.isoformat() if t.date else None,
                "description": t.description,
                "amount": float(t.amount),
                "category": t.bucket.name if t.bucket else "Uncategorized",
                "account": t.account.name if t.account else "Unknown",
                "type": "income" if t.amount > 0 else "expense",
                "notes": t.notes
            }
            for t in transactions
        ]
        
        # Create a generator-like stream for JSON
        json_str = json.dumps(data, indent=2)
        return StreamingResponse(
            io.StringIO(json_str),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=transactions_export_{datetime.now().strftime('%Y%m%d')}.json"}
        )
        
    else: # CSV
        stream = io.StringIO()
        writer = csv.writer(stream)
        
        # Write header
        writer.writerow(["Date", "Description", "Amount", "Type", "Category", "Account", "Notes"])
        
        # Write data
        for t in transactions:
            writer.writerow([
                t.date,
                t.description,
                t.amount,
                "Income" if t.amount > 0 else "Expense",
                t.bucket.name if t.bucket else "Uncategorized",
                t.account.name if t.account else "Unknown",
                t.notes or ""
            ])
            
        stream.seek(0)
        return StreamingResponse(
            iter([stream.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=transactions_export_{datetime.now().strftime('%Y%m%d')}.csv"}
        )

@router.get("/net-worth")
def export_net_worth(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    # Fetch history
    query = db.query(models.NetWorthSnapshot)
    
    if start_date:
        query = query.filter(models.NetWorthSnapshot.date >= start_date)
    if end_date:
        query = query.filter(models.NetWorthSnapshot.date <= end_date)
        
    snapshots = query.order_by(models.NetWorthSnapshot.date.desc()).all()
    
    stream = io.StringIO()
    writer = csv.writer(stream)
    
    # Header
    writer.writerow(["Date", "Total Net Worth", "Total Assets", "Total Liabilities"])
    
    # Data
    for s in snapshots:
        writer.writerow([
            s.date,
            s.net_worth,
            s.total_assets,
            s.total_liabilities
        ])
        
    stream.seek(0)
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=net_worth_history_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.get("/report/pdf")
def export_report_pdf(
    start_date: str = Query(..., description="ISO Date string"),
    end_date: str = Query(..., description="ISO Date string"),
    spender: str = Query(default="Combined"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Generate a professional PDF financial report (One Pager).
    Includes: Large PNG Logo, Blue Bar Headers, Donut Charts, Compact Layout.
    Fits on single page.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as ReportLabImage
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.piecharts import Pie
    import io
    import os
    
    # Brand colors
    BRAND_PRIMARY = colors.HexColor('#5D5DFF')  # Electric Indigo
    BRAND_SUCCESS = colors.HexColor('#34D399')  # Emerald
    BRAND_WARNING = colors.HexColor('#FB923C')  # Orange
    BRAND_ERROR = colors.HexColor('#EF4444')    # Red
    TEXT_PRIMARY = colors.HexColor('#191B18')
    TEXT_MUTED = colors.HexColor('#666666')
    BACKGROUND_SURFACE = colors.HexColor('#F8FAFC')
    BORDER_COLOR = colors.HexColor('#E2E8F0')
    WHITE = colors.white
    
    # Chart colors (Brand Palette + Distinct additions)
    CHART_COLORS = [
        BRAND_PRIMARY,    # #5D5DFF
        BRAND_SUCCESS,    # #34D399
        BRAND_WARNING,    # #FB923C
        BRAND_ERROR,      # #EF4444
        colors.HexColor('#8B5CF6'),  # Violet
        colors.HexColor('#EC4899'),  # Pink
        colors.HexColor('#06B6D4'),  # Cyan
        colors.HexColor('#F59E0B'),  # Amber
        colors.HexColor('#6366F1'),  # Indigo-500
        colors.HexColor('#10B981'),  # Emerald-500
    ]
    
    try:
        s_date = datetime.fromisoformat(start_date)
        e_date = datetime.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Fetch transactions - EXCLUDE TRANSFERS
    transactions = db.query(models.Transaction).join(
        models.BudgetBucket, models.Transaction.bucket_id == models.BudgetBucket.id, isouter=True
    ).filter(
        models.Transaction.user_id == current_user.id,
        models.Transaction.date >= s_date,
        models.Transaction.date <= e_date
    )
    
    # Exclude transfers
    transactions = transactions.filter(
        ~models.BudgetBucket.name.ilike('%transfer%')
    )
    
    if spender != "Combined":
        transactions = transactions.filter(models.Transaction.spender == spender)
    
    transactions = transactions.all()
    
    # === PROCESS DATA ===
    total_income = 0
    total_expenses = 0
    total_invested = 0
    
    expense_categories = {}
    income_categories = {}
    
    for t in transactions:
        cat_name = t.bucket.name if t.bucket else "Uncategorized"
        is_investment = "investment" in cat_name.lower()
        
        if t.amount > 0:
            # Income
            total_income += t.amount
            income_categories[cat_name] = income_categories.get(cat_name, 0) + t.amount
        else:
            # Outflow
            abs_amount = abs(t.amount)
            if is_investment:
                total_invested += abs_amount
            else:
                total_expenses += abs_amount
                expense_categories[cat_name] = expense_categories.get(cat_name, 0) + abs_amount
                
    net_savings = total_income - total_expenses
    savings_rate = (net_savings / total_income * 100) if total_income > 0 else 0
    
    # Create PDF - Tighter margins for One Pager
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*cm, bottomMargin=0.5*cm, leftMargin=1.0*cm, rightMargin=1.0*cm)
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=28, spaceAfter=12, textColor=TEXT_PRIMARY, alignment=0) 
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=TEXT_MUTED, alignment=0)
    
    # Blue Bar Section Header Style
    def get_section_header_table(text):
        p = Paragraph(text, ParagraphStyle('SectionHeader', parent=styles['Normal'], fontSize=12, textColor=WHITE, fontName='Helvetica-Bold'))
        t = Table([[p]], colWidths=[7.3*inch]) # SAFE WIDTH
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), BRAND_PRIMARY),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        return t

    elements = []
    
    # === HEADER: Large Logo + Title (Side by Side) ===
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logo.png")
    
    logo_img = ""
    if os.path.exists(logo_path):
        try:
             # Larger logo: width=3.5inch
            logo_img = ReportLabImage(logo_path, width=3.5*inch, height=1.2*inch, kind='proportional')
            logo_img.hAlign = 'LEFT'
        except Exception: pass
    
    title_text = "Financial Report"
    period_text = f"{s_date.strftime('%d %b %Y')} - {e_date.strftime('%d %b %Y')}"
    generated_text = f"Generated: {datetime.now().strftime('%d %b %Y')}"
    
    # Meta info stack
    meta_elements = [
        Paragraph(title_text, title_style),
        Paragraph(period_text, subtitle_style),
    ]
    if spender != "Combined":
        meta_elements.append(Paragraph(f"Spender: {spender}", subtitle_style))
    meta_elements.append(Paragraph(generated_text, subtitle_style))
    
    # Header Table: Logo Left | Meta Right
    # Reduced to 7.3 total
    header_table = Table([[logo_img if logo_img else "", meta_elements]], colWidths=[3.5*inch, 3.8*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # === HIGHLIGHTS / SUMMARY GRID ===
    # A single row of 4 boxes? Or 2x2. 1x4 is cleaner for One Pager top.
    
    def create_stat_box(label, value, color=TEXT_PRIMARY):
        return [
            Paragraph(label, ParagraphStyle('StatLabel', parent=styles['Normal'], fontSize=9, textColor=TEXT_MUTED, alignment=1)),
            Paragraph(value, ParagraphStyle('StatValue', parent=styles['Normal'], fontSize=14, textColor=color, fontName='Helvetica-Bold', alignment=1))
        ]
        
    stats_data = [[
        create_stat_box("Total Income", f"${total_income:,.0f}", BRAND_SUCCESS),
        create_stat_box("Total Expenses", f"${total_expenses:,.0f}", BRAND_ERROR),
        create_stat_box("Net Savings", f"${net_savings:,.0f}", BRAND_SUCCESS if net_savings >= 0 else BRAND_ERROR),
        create_stat_box("Savings Rate", f"{savings_rate:.1f}%", BRAND_PRIMARY)
    ]]
    
    # Reduced to 7.2 total (1.8 * 4) to be very safe
    stats_table = Table(stats_data, colWidths=[1.8*inch]*4)
    stats_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (0,0), 0.5, BORDER_COLOR),
        ('BOX', (1,0), (1,0), 0.5, BORDER_COLOR),
        ('BOX', (2,0), (2,0), 0.5, BORDER_COLOR),
        ('BOX', (3,0), (3,0), 0.5, BORDER_COLOR),
        ('BACKGROUND', (0,0), (-1,-1), BACKGROUND_SURFACE),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # === SECTION: Spending & Income (Blue Bar) ===
    elements.append(get_section_header_table("Cash Flow Allocation"))
    elements.append(Spacer(1, 0.2*inch))

    # === DONUT CHARTS ===
    
    # Sort data
    exp_sorted = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)[:6]
    inc_sorted = sorted(income_categories.items(), key=lambda x: x[1], reverse=True)[:6]
    
    # Filter small values to prevent chart crashes
    exp_values = [float(x[1]) for x in exp_sorted if float(x[1]) > 0.01]
    inc_values = [float(x[1]) for x in inc_sorted if float(x[1]) > 0.01]
    
    def create_donut_drawing(values, colors_list, width=125, height=125):
        if not values: return None
        d = Drawing(width, height)
        pie = Pie()
        pie.x = 10
        pie.y = 10
        pie.width = width - 20
        pie.height = height - 20
        # pie.innerRadius is NOT supported in standard ReportLab Pie. 
        # We simulate a donut by drawing a white circle on top.
        
        pie.data = values
        pie.labels = None
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.white
        for i in range(len(values)):
            pie.slices[i].fillColor = colors_list[i % len(colors_list)]
        d.add(pie)
        
        # Draw white circle for Donut hole
        from reportlab.graphics.shapes import Circle
        cx = width / 2
        cy = height / 2
        radius = (width - 20) * 0.35
        c = Circle(cx, cy, radius)
        c.fillColor = colors.white
        c.strokeColor = colors.white
        d.add(c)
        
        return d
    
    inc_colors = CHART_COLORS[3:] + CHART_COLORS[:3]
    exp_drawing = create_donut_drawing(exp_values, CHART_COLORS)
    inc_drawing = create_donut_drawing(inc_values, inc_colors)

    # Simplified Side-Legend
    def create_side_legend(sorted_items, color_palette):
        if not sorted_items: return Paragraph("No data", subtitle_style)
        l_data = []
        total = sum([x[1] for x in sorted_items])
        for i, (label, value) in enumerate(sorted_items):
            color = color_palette[i % len(color_palette)]
            pct = (value/total*100) if total > 0 else 0
            label_trunc = (label[:15] + '.') if len(label) > 15 else label
            l_data.append(["", f"{label_trunc}", f"{pct:.0f}%"])
            
        t = Table(l_data, colWidths=[0.15*inch, 1.15*inch, 0.4*inch])
        styles_list = [
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), TEXT_MUTED),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'), # % Right align
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ]
        for i in range(len(l_data)):
            color = color_palette[i % len(color_palette)]
            styles_list.append(('BACKGROUND', (0, i), (0, i), color))
            styles_list.append(('BOTTOMPADDING', (0, i), (-1, i), 3)) # More padding
            styles_list.append(('TOPPADDING', (0, i), (-1, i), 3))
            # Add white line below to create "gap" between colored boxes
            styles_list.append(('LINEBELOW', (0, i), (-1, i), 3, colors.white))
            
        t.setStyle(TableStyle(styles_list))
        return t

    exp_legend_side = create_side_legend(exp_sorted, CHART_COLORS)
    inc_legend_side = create_side_legend(inc_sorted, inc_colors)
    
    # Layout: Chart | Legend || Chart | Legend
    chart_row = [
        [Paragraph("Expenses", subtitle_style), exp_drawing if exp_drawing else "", exp_legend_side],
        [Paragraph("Income", subtitle_style), inc_drawing if inc_drawing else "", inc_legend_side]
    ]
    
    # Actually, let's make nested tables for each group: [Title] / [Chart | Legend]
    def create_chart_group(title, drawing, legend):
        # Handle None drawing safely
        content = [
            [drawing if drawing else Paragraph("No Data", subtitle_style), legend]
        ]
        t = Table(content, colWidths=[1.8*inch, 1.8*inch])
        t.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        
        container = Table([[Paragraph(title, subtitle_style)], [t]], colWidths=[3.6*inch])
        container.setStyle(TableStyle([('LEFTPADDING', (0,0), (-1,-1), 0)]))
        return container
        
    exp_group = create_chart_group("Expenses Breakdown", exp_drawing, exp_legend_side)
    inc_group = create_chart_group("Income Breakdown", inc_drawing, inc_legend_side)
    
    main_charts = Table([[exp_group, inc_group]], colWidths=[3.65*inch, 3.65*inch])
    main_charts.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    
    elements.append(main_charts)
    elements.append(Spacer(1, 0.2*inch))
    
    # === SECTION: Detailed Data (Blue Bar) ===
    elements.append(get_section_header_table("Top Categories"))
    elements.append(Spacer(1, 0.2*inch))
    
    # Detailed Tables (Top 8)
    exp_full_sorted = sorted(expense_categories.items(), key=lambda x: x[1], reverse=True)[:8]
    inc_full_sorted = sorted(income_categories.items(), key=lambda x: x[1], reverse=True)[:8]
    
    exp_data = [["Category", "Amount", "%"]]
    for cat_name, amount in exp_full_sorted:
        pct = (amount / total_expenses * 100) if total_expenses > 0 else 0
        exp_data.append([cat_name[:22], f"${amount:,.0f}", f"{pct:.0f}%"])
        
    inc_data = [["Category", "Amount", "%"]]
    for cat_name, amount in inc_full_sorted:
        pct = (amount / total_income * 100) if total_income > 0 else 0
        inc_data.append([cat_name[:22], f"${amount:,.0f}", f"{pct:.0f}%"])
        
    while len(exp_data) < len(inc_data): exp_data.append(["", "", ""])
    while len(inc_data) < len(exp_data): inc_data.append(["", "", ""])
    
    # Clean Table Style (No zebra, just lines)
    clean_ts = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8), # Smaller font
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_MUTED),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BORDER_COLOR), # Header line
        ('LINEBELOW', (0, 1), (-1, -1), 0.5, BORDER_COLOR), # Row lines
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    
    exp_table = Table(exp_data, colWidths=[2.0*inch, 1.0*inch, 0.5*inch])
    exp_table.setStyle(TableStyle(clean_ts))
    
    inc_table = Table(inc_data, colWidths=[2.0*inch, 1.0*inch, 0.5*inch])
    inc_table.setStyle(TableStyle(clean_ts))
    
    tables_layout = Table([[exp_table, Spacer(0.2*inch, 0), inc_table]], colWidths=[3.5*inch, 0.2*inch, 3.5*inch])
    tables_layout.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    
    elements.append(tables_layout)
    
    # Footer
    elements.append(Spacer(1, 0.4*inch))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=TEXT_MUTED, alignment=1)
    elements.append(Paragraph("Generated by DollarData • dollardata.au", footer_style))
    
    # Build PDF
    try:
        doc.build(elements)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")
        
    buffer.seek(0)
    
    filename = f"DollarData_Report_{s_date.strftime('%Y%m')}_to_{e_date.strftime('%Y%m')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/report/excel")
def export_report_excel(
    start_date: str = Query(..., description="ISO Date string"),
    end_date: str = Query(..., description="ISO Date string"),
    spender: str = Query(default="Combined"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Generate an Excel workbook with multiple sheets.
    Sheets: Summary, Transactions, Category Breakdown
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    try:
        s_date = datetime.fromisoformat(start_date)
        e_date = datetime.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Fetch data
    transactions = db.query(models.Transaction).filter(
        models.Transaction.user_id == current_user.id,
        models.Transaction.date >= s_date,
        models.Transaction.date <= e_date
    )
    
    if spender != "Combined":
        transactions = transactions.filter(models.Transaction.spender == spender)
    
    transactions = transactions.order_by(models.Transaction.date.desc()).all()
    
    # Calculate stats
    total_income = sum(t.amount for t in transactions if t.amount > 0)
    total_expenses = sum(abs(t.amount) for t in transactions if t.amount < 0)
    net_savings = total_income - total_expenses
    
    # Category breakdown
    category_totals = {}
    for t in transactions:
        if t.amount < 0:
            cat_name = t.bucket.name if t.bucket else "Uncategorized"
            category_totals[cat_name] = category_totals.get(cat_name, 0) + abs(t.amount)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5D5DFF", end_color="5D5DFF", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    currency_format = '$#,##0.00'
    
    # ===== Summary Sheet =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = "DollarData Financial Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Period: {s_date.strftime('%B %d, %Y')} - {e_date.strftime('%B %d, %Y')}"
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if spender != "Combined":
        ws_summary['A4'] = f"Spender: {spender}"
    
    summary_data = [
        ("Metric", "Value"),
        ("Total Income", total_income),
        ("Total Expenses", total_expenses),
        ("Net Savings", net_savings),
        ("Savings Rate", f"{(net_savings/total_income*100):.1f}%" if total_income > 0 else "N/A"),
        ("Transaction Count", len(transactions)),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, start=6):
        ws_summary[f'A{row_idx}'] = label
        ws_summary[f'B{row_idx}'] = value
        if row_idx == 6:
            ws_summary[f'A{row_idx}'].font = header_font
            ws_summary[f'A{row_idx}'].fill = header_fill
            ws_summary[f'B{row_idx}'].font = header_font
            ws_summary[f'B{row_idx}'].fill = header_fill
        elif isinstance(value, (int, float)) and row_idx > 6:
            ws_summary[f'B{row_idx}'].number_format = currency_format
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    # ===== Transactions Sheet =====
    ws_txn = wb.create_sheet("Transactions")
    
    headers = ["Date", "Description", "Amount", "Type", "Category", "Spender", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws_txn.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, t in enumerate(transactions, start=2):
        ws_txn.cell(row=row_idx, column=1, value=t.date.strftime("%Y-%m-%d") if t.date else "")
        ws_txn.cell(row=row_idx, column=2, value=t.description)
        amount_cell = ws_txn.cell(row=row_idx, column=3, value=t.amount)
        amount_cell.number_format = currency_format
        ws_txn.cell(row=row_idx, column=4, value="Income" if t.amount > 0 else "Expense")
        ws_txn.cell(row=row_idx, column=5, value=t.bucket.name if t.bucket else "Uncategorized")
        ws_txn.cell(row=row_idx, column=6, value=t.spender or "")
        ws_txn.cell(row=row_idx, column=7, value=t.notes or "")
    
    # Auto-size columns
    for col in range(1, 8):
        ws_txn.column_dimensions[get_column_letter(col)].width = 15 if col != 2 else 40
    
    # ===== Category Breakdown Sheet =====
    ws_cat = wb.create_sheet("Category Breakdown")
    
    cat_headers = ["Category", "Amount", "% of Total"]
    for col, header in enumerate(cat_headers, start=1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    sorted_cats = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
    for row_idx, (cat_name, amount) in enumerate(sorted_cats, start=2):
        ws_cat.cell(row=row_idx, column=1, value=cat_name)
        amount_cell = ws_cat.cell(row=row_idx, column=2, value=amount)
        amount_cell.number_format = currency_format
        pct = (amount / total_expenses * 100) if total_expenses > 0 else 0
        ws_cat.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
    
    ws_cat.column_dimensions['A'].width = 30
    ws_cat.column_dimensions['B'].width = 15
    ws_cat.column_dimensions['C'].width = 12
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"DollarData_Report_{s_date.strftime('%Y%m')}_to_{e_date.strftime('%Y%m')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

